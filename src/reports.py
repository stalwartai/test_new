"""
Report Generator — creates grouped Excel reports.
Each row = one story, with all sources listed.
"""
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Config
from .database import Database

logger = logging.getLogger('news_tracker')


class ReportGenerator:
    def __init__(self):
        self.db = Database()
        self.output_dir = Config.OUTPUT_FOLDER

    def generate_daily_report(self, days=7):
        """Generate an Excel report with grouped stories."""
        try:
            stories = self.db.get_stories_grouped(days=days)

            if not stories:
                logger.warning("No stories to report")
                return None

            # Build data rows
            rows = []
            for story in stories:
                source_links = []
                for s in story['sources']:
                    source_links.append(s['url'])

                rows.append({
                    'Date': story['published'],
                    'Person': story['person'],
                    'Story Headline': story['headline'],
                    'Category': story['category'],
                    'Total Sources': story['source_count'],
                    'Source Names': story['source_names'],
                    'Languages': story['languages'],
                    'Links': '\n'.join(source_links[:10]),  # Max 10 links per cell
                })

            df = pd.DataFrame(rows)

            # Generate filename
            timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
            filename = f"news_report_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            # Write to Excel
            df.to_excel(filepath, index=False, sheet_name='News Stories')

            # Format the Excel file
            self._format_excel(filepath, df)

            logger.info(f"Report generated: {filepath}")
            return filepath

        except Exception as e:
            logger.exception(f"Error generating report: {e}")
            return None

    def _format_excel(self, filepath, df):
        """Apply professional formatting to the Excel file."""
        try:
            wb = load_workbook(filepath)
            ws = wb.active

            # Header styling
            header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1B2838', end_color='1B2838', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Style headers
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Style data rows
            alt_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
            data_font = Font(name='Calibri', size=10)

            for row_num in range(2, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    if row_num % 2 == 0:
                        cell.fill = alt_fill

            # Column widths
            column_widths = {
                'A': 18,   # Date
                'B': 16,   # Person
                'C': 50,   # Headline
                'D': 14,   # Category
                'E': 12,   # Total Sources
                'F': 35,   # Source Names
                'G': 12,   # Languages
                'H': 50,   # Links
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Freeze top row
            ws.freeze_panes = 'A2'

            wb.save(filepath)
        except Exception as e:
            logger.error(f"Error formatting Excel: {e}")

    def generate_summary(self, days=7):
        """Generate a text summary of statistics."""
        stats = self.db.get_statistics(days=days)
        summary = f"""
╔══════════════════════════════════════╗
║     NEWS TRACKER SUMMARY ({days}d)       ║
╠══════════════════════════════════════╣
║ Total Articles: {stats['total_articles']:>5}               ║
║ Total Stories:  {stats['total_stories']:>5}               ║
║ Modi Coverage:  {stats['modi_count']:>5}               ║
║ Patil Coverage: {stats['patil_count']:>5}               ║
║ Unique Channels:{stats['unique_channels']:>5}               ║
║ Languages:      {', '.join(stats['languages']):>18} ║
╚══════════════════════════════════════╝
"""
        return summary
